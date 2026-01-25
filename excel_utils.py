# File: excel_utils.py
import os
import io
from openpyxl import load_workbook
from xhtml2pdf import pisa 

# 1. HÀM TẠO EXCEL (Giữ nguyên)
def create_excel_report(date, supervisor, output, notes):
    filename = 'mau_bao_cao.xlsx'
    if not os.path.exists(filename):
        return None, "Lỗi: Không tìm thấy file mẫu 'mau_bao_cao.xlsx'!"
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        # Điền dữ liệu
        ws['A5'] = date
        ws['C4'] = supervisor
        ws['C5'] = output
        ws['B8'] = notes
        
        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)
        return output_stream, None
    except Exception as e:
        return None, f"Lỗi Excel: {str(e)}"

# 2. HÀM TẠO PDF (Đã sửa giao diện giống Excel)
def create_pdf_report(date, supervisor, output, notes):
    try:
        # Tèo dùng HTML để vẽ lại cái bảng y chang Excel
        # Anh có thể chỉnh sửa độ rộng (width), chiều cao (height) ở đây
        html_content = f"""
        <html>
        <head>
            <style>
                body {{ 
                    font-family: Arial, sans-serif; 
                    padding: 40px; 
                    font-size: 14px;
                }}
                /* Tiêu đề báo cáo */
                h1 {{ 
                    text-align: center; 
                    text-transform: uppercase; 
                    margin-bottom: 30px;
                    color: #000;
                }}
                
                /* Cái bảng mô phỏng Excel */
                table {{
                    width: 100%;
                    border-collapse: collapse; /* Gom đường viền lại cho mỏng */
                    margin-bottom: 20px;
                }}
                
                /* Kẻ khung cho từng ô */
                td, th {{
                    border: 1px solid black; /* Viền đen 1px */
                    padding: 10px; /* Khoảng cách chữ cách viền */
                    vertical-align: middle;
                }}
                
                /* Cột tiêu đề (Bên trái) */
                .label-col {{
                    background-color: #f0f0f0; /* Tô màu xám nhẹ giống header Excel */
                    font-weight: bold;
                    width: 30%; /* Chiếm 30% chiều rộng */
                }}
                
                /* Cột dữ liệu (Bên phải) */
                .data-col {{
                    width: 70%;
                    font-weight: bold;
                    color: #333;
                }}
                
                /* Phần ghi chú (Merge cells) */
                .note-header {{
                    background-color: #f0f0f0;
                    font-weight: bold;
                    text-align: left;
                }}
                .note-content {{
                    height: 100px; /* Chiều cao cố định cho ô ghi chú */
                    vertical-align: top; /* Chữ nằm trên cùng */
                }}
                
                /* Phần chữ ký */
                .footer {{
                    margin-top: 40px;
                    text-align: right;
                    width: 100%;
                }}
            </style>
        </head>
        <body>
            <h1>BÁO CÁO SẢN XUẤT</h1>

            <table>
                <tr>
                    <td class="label-col">Ngày báo cáo</td>
                    <td class="data-col">{date}</td>
                </tr>
                <tr>
                    <td class="label-col">Người phụ trách</td>
                    <td class="data-col">{supervisor}</td>
                </tr>
                <tr>
                    <td class="label-col">Tổng sản lượng (Tấn)</td>
                    <td class="data-col">{output}</td>
                </tr>
                
                <tr>
                    <td colspan="2" class="note-header">Ghi chú / Sự cố trong ca trực:</td>
                </tr>
                <tr>
                    <td colspan="2" class="note-content">
                        {notes.replace(chr(10), '<br>')}
                    </td>
                </tr>
            </table>

            <div class="footer">
                <p><i>Ngày......tháng......năm......</i></p>
                <p><b>Người lập biểu</b></p>
                <br><br><br>
                <p>{supervisor}</p>
            </div>
        </body>
        </html>
        """
        
        pdf_out = io.BytesIO()
        pisa_status = pisa.CreatePDF(io.StringIO(html_content), dest=pdf_out)
        
        if pisa_status.err:
            return None, "Lỗi thư viện tạo PDF"
            
        pdf_out.seek(0)
        return pdf_out, None

    except Exception as e:
        return None, f"Lỗi PDF: {str(e)}"